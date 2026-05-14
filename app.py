from flask import Flask, jsonify, request, render_template
from pathlib import Path
import openpyxl
import json
import sqlite3
import csv
import ijson

app = Flask(__name__)
TRACKER   = Path(__file__).parent / "tracker.xlsx"
JSON_FILE = Path(__file__).parent / "ettevotja_rekvisiidid__yldandmed.json"
CSV_FILE      = Path(__file__).parent / "4.2025_aruannete_elemendid_kuni_30042026.csv"
CSV_META_FILE = Path(__file__).parent / "1.aruannete_yldandmed_kuni_30042026.csv"
ISIKUD_FILE   = Path(__file__).parent / "ettevotja_rekvisiidid__kaardile_kantud_isikud.json"
DB_FILE   = Path(__file__).parent / "ariregister_cache.db"


# ── SQLite cache for the 4 GB JSON ──────────────────────────────────────────

def _get_db():
    conn = sqlite3.connect(str(DB_FILE))
    conn.row_factory = sqlite3.Row
    return conn


WANT_LABELS = {
    "Müügitulu":                                        "myygitulu",
    "Ärikasum (kahjum)":                               "arikasum",
    "Aruandeaasta kasum (kahjum)":                     "kasum",
    "Töötajate keskmine arv taandatuna täistööajale":  "tootajad_arv",
    "Varad":                                           "varad",
    "Tööjõukulud":                                     "toojoukulud",
}


def _load_financials():
    """Build {registrikood: {field: value}} using latest report year per company."""
    print("Loading report metadata…")
    # Step 1: for each registrikood keep the report_id of the latest aruandeaasta
    latest = {}  # registrikood -> (aruandeaasta, report_id)
    with open(str(CSV_META_FILE), "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            kood  = row["registrikood"].strip()
            year  = row["aruandeaasta"].strip()
            rid   = row["report_id"].strip()
            if not kood or not year or not rid:
                continue
            if kood not in latest or year > latest[kood][0]:
                latest[kood] = (year, rid)

    # report_id -> registrikood (inverted for fast lookup)
    rid_to_kood = {v[1]: k for k, v in latest.items()}
    print(f"  {len(latest)} companies with reports, {len(rid_to_kood)} report_ids indexed.")

    # Step 2: read financial elements, only for the latest reports
    print("Loading financial elements…")
    data = {}  # registrikood -> {field: value}
    with open(str(CSV_FILE), "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            label = row["elemendi_label"]
            if label not in WANT_LABELS:
                continue
            rid = row["report_id"].strip()
            if rid not in rid_to_kood:
                continue
            kood = rid_to_kood[rid]
            val_str = row["vaartus"].replace(",", ".").strip()
            try:
                val = float(val_str) if val_str else None
            except ValueError:
                val = None
            if kood not in data:
                data[kood] = {}
            data[kood][WANT_LABELS[label]] = val

    print(f"  Loaded financials for {len(data)} companies.")
    return data


def _load_board_members():
    """Return {ariregistri_kood: 'Eesnimi Perenimi, ...'} for active JUHL members."""
    print("Loading board members…")
    data = {}
    with open(str(ISIKUD_FILE), "r", encoding="utf-8") as f:
        for company in ijson.items(f, "item"):
            kood = company.get("ariregistri_kood")
            isikud = company.get("kaardile_kantud_isikud") or []
            names = []
            for p in isikud:
                if p.get("isiku_roll") == "JUHL" and not p.get("lopp_kpv"):
                    eesnimi = p.get("eesnimi") or ""
                    perenimi = p.get("nimi_arinimi") or ""
                    full = f"{eesnimi} {perenimi}".strip()
                    if full:
                        names.append(full)
            if names:
                data[kood] = ", ".join(names)
    print(f"  Loaded board members for {len(data)} companies.")
    return data


def build_cache():
    """Stream-parse the JSON and write rows into SQLite. Run once."""
    conn = _get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            kood             INTEGER PRIMARY KEY,
            nimi             TEXT,
            staatus          TEXT,
            vorm             TEXT,
            piirkond         TEXT,
            aadress          TEXT,
            kapital          REAL,
            valuuta          TEXT,
            reg_kpv          TEXT,
            emtak_kood       TEXT,
            emtak_nimi       TEXT,
            email            TEXT,
            telefon          TEXT,
            veebileht        TEXT,
            tootajad         INTEGER,
            myygitulu        REAL,
            arikasum         REAL,
            kasum            REAL,
            tootajad_arv     REAL,
            varad            REAL,
            toojoukulud      REAL,
            juhatuse_liige   TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_staatus  ON companies(staatus)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_vorm     ON companies(vorm)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_piirkond ON companies(piirkond)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_emtak    ON companies(emtak_kood)")
    conn.commit()

    count = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    if count > 0:
        conn.close()
        return

    financials    = _load_financials()
    board_members = _load_board_members()

    print("Building ariregister cache from JSON (one-time, may take a few minutes)…")
    with open(str(JSON_FILE), "r", encoding="utf-8") as f:
        rows = []
        for company in ijson.items(f, "item"):
            y = company.get("yldandmed", {})

            aadressid = y.get("aadressid") or []
            adr  = aadressid[0].get("aadress_ads__ads_normaliseeritud_taisaadress", "") if aadressid else ""
            linn = aadressid[0].get("ehak_nimetus", "") if aadressid else ""

            kapitalid = y.get("kapitalid") or []
            kapital = float(kapitalid[0].get("kapitali_suurus", 0) or 0) if kapitalid else 0
            valuuta = kapitalid[0].get("kapitali_valuuta_tekstina", "") if kapitalid else ""

            tegevusalad = y.get("teatatud_tegevusalad") or []
            pohiteg    = next((t for t in tegevusalad if t.get("on_pohitegevusala")), tegevusalad[0] if tegevusalad else {})
            emtak_kood = pohiteg.get("emtak_kood", "")
            emtak_nimi = pohiteg.get("emtak_tekstina", "")

            sidevahendid = y.get("sidevahendid") or []
            email = telefon = veebileht = ""
            for sv in sidevahendid:
                liik = sv.get("liik", "")
                sisu = sv.get("sisu", "")
                if liik == "EMAIL" and not email:          email     = sisu
                elif liik in ("MOB", "TEL") and not telefon: telefon = sisu
                elif liik == "WWW" and not veebileht:     veebileht = sisu

            aruanded = y.get("info_majandusaasta_aruannetest") or []
            tootajad = None
            if aruanded:
                raw = aruanded[-1].get("tootajate_arv")
                try:    tootajad = int(raw) if raw not in (None, "") else None
                except: tootajad = None

            kood = company.get("ariregistri_kood")
            fin  = financials.get(str(kood), {})

            # Prefer employee count from financial report (average FTE), fall back to JSON
            if fin.get("tootajad_arv") is not None:
                tootajad = int(round(fin["tootajad_arv"]))

            rows.append((
                kood,
                company.get("nimi", ""),
                y.get("staatus_tekstina", ""),
                y.get("oiguslik_vorm_tekstina", ""),
                linn,
                adr,
                kapital,
                valuuta,
                y.get("esmaregistreerimise_kpv", ""),
                emtak_kood,
                emtak_nimi,
                email,
                telefon,
                veebileht,
                tootajad,
                fin.get("myygitulu"),
                fin.get("arikasum"),
                fin.get("kasum"),
                fin.get("tootajad_arv"),
                fin.get("varad"),
                fin.get("toojoukulud"),
                board_members.get(kood),
            ))

            if len(rows) >= 2000:
                conn.executemany("INSERT OR IGNORE INTO companies VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", rows)
                conn.commit()
                rows = []

        if rows:
            conn.executemany("INSERT OR IGNORE INTO companies VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", rows)
            conn.commit()

    print("Cache built.")
    conn.close()


ALLOWED_SORT = {
    "nimi", "staatus", "vorm", "piirkond", "reg_kpv",
    "kapital", "tootajad", "myygitulu", "arikasum", "kasum", "varad", "toojoukulud"
}

def query_companies(search="", staatus="", vorm="", piirkond="", emtak="",
                    rev_min=None, rev_max=None, emp_min=None, emp_max=None,
                    profit_min=None, profit_max=None,
                    limit=200, offset=0, sort_key="nimi", sort_dir="asc"):
    conn = _get_db()
    clauses, params = [], []

    if search:
        clauses.append("(nimi LIKE ? OR aadress LIKE ? OR emtak_nimi LIKE ? OR email LIKE ?)")
        like = f"%{search}%"
        params += [like, like, like, like]
    if staatus:
        clauses.append("staatus = ?"); params.append(staatus)
    if vorm:
        clauses.append("vorm = ?"); params.append(vorm)
    if piirkond:
        clauses.append("piirkond LIKE ?"); params.append(f"%{piirkond}%")
    if emtak:
        clauses.append("emtak_kood LIKE ?"); params.append(f"{emtak}%")
    if rev_min is not None:
        clauses.append("myygitulu >= ?"); params.append(rev_min)
    if rev_max is not None:
        clauses.append("myygitulu <= ?"); params.append(rev_max)
    if emp_min is not None:
        clauses.append("tootajad_arv >= ?"); params.append(emp_min)
    if emp_max is not None:
        clauses.append("tootajad_arv <= ?"); params.append(emp_max)
    if profit_min is not None:
        clauses.append("arikasum >= ?"); params.append(profit_min)
    if profit_max is not None:
        clauses.append("arikasum <= ?"); params.append(profit_max)

    if sort_key not in ALLOWED_SORT:
        sort_key = "nimi"
    order = f"{sort_key} {'DESC' if sort_dir == 'desc' else 'ASC'} NULLS LAST"

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    total = conn.execute(f"SELECT COUNT(*) FROM companies {where}", params).fetchone()[0]
    rows  = conn.execute(
        f"SELECT * FROM companies {where} ORDER BY {order} LIMIT ? OFFSET ?",
        params + [limit, offset]
    ).fetchall()
    conn.close()
    return total, [dict(r) for r in rows]


def get_filter_options():
    conn = _get_db()
    statuses = [r[0] for r in conn.execute("SELECT DISTINCT staatus FROM companies WHERE staatus != '' ORDER BY staatus").fetchall()]
    vormed   = [r[0] for r in conn.execute("SELECT DISTINCT vorm    FROM companies WHERE vorm    != '' ORDER BY vorm").fetchall()]
    conn.close()
    return statuses, vormed

STATUSES = ["New", "Contacted", "Interested", "Not Interested", "Follow Up", "Closed Deal", "Manually Declined"]

FIELD_MAP = {
    'field':         'Field',
    'name':          'Name',
    'url':           'LinkedIn URL',
    'industry':      'Industry',
    'location':      'Location',
    'description':   'Description',
    'score':         'Automation Interest Score (1-10)',
    'needs':         'Potential Needs',
    'contact_role':  'Recommended Contact Role',
    'contact_person':'Contact Person',
    'status':        'Status',
}


def load_companies():
    wb = openpyxl.load_workbook(TRACKER)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    companies = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        r = dict(zip(headers, row))
        score_raw = r.get("Automation Interest Score (1-10)")
        try:
            score = int(score_raw) if score_raw not in (None, '', 'Not set') else 0
        except (ValueError, TypeError):
            score = 0
        companies.append({
            "row":          i,
            "id":           int(r.get("#") or i - 1),
            "field":        r.get("Field") or "",
            "name":         r.get("Name") or "",
            "url":          r.get("LinkedIn URL") or "",
            "industry":     r.get("Industry") or "",
            "location":     r.get("Location") or "",
            "description":  r.get("Description") or "",
            "score":        score,
            "needs":        r.get("Potential Needs") or "",
            "contact_role": r.get("Recommended Contact Role") or "",
            "contact_person": r.get("Contact Person") or "",
            "status":       r.get("Status") or "",
        })
    return companies


def save_field(row_num: int, col_name: str, value):
    wb = openpyxl.load_workbook(TRACKER)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_idx = headers.index(col_name) + 1
    ws.cell(row=row_num, column=col_idx, value=value)
    wb.save(TRACKER)


@app.route("/")
def index():
    return render_template("index.html", statuses=STATUSES)


@app.route("/ariregister")
def ariregister():
    statuses, vormed = get_filter_options()
    return render_template("ariregister.html",
                           statuses=statuses,
                           vormed=vormed)


@app.route("/api/ariregister")
def api_ariregister():
    search   = request.args.get("search", "").strip()
    staatus  = request.args.get("staatus", "").strip()
    vorm     = request.args.get("vorm", "").strip()
    piirkond = request.args.get("piirkond", "").strip()
    emtak    = request.args.get("emtak", "").strip()
    limit    = min(int(request.args.get("limit", 200)), 500)
    offset   = int(request.args.get("offset", 0))
    sort_key = request.args.get("sort_key", "nimi").strip()
    sort_dir = request.args.get("sort_dir", "asc").strip()

    def _float_or_none(key):
        v = request.args.get(key, "").strip()
        try: return float(v) if v else None
        except ValueError: return None

    rev_min = _float_or_none("rev_min")
    rev_max = _float_or_none("rev_max")
    emp_min    = _float_or_none("emp_min")
    emp_max    = _float_or_none("emp_max")
    profit_min = _float_or_none("profit_min")
    profit_max = _float_or_none("profit_max")

    total, rows = query_companies(search, staatus, vorm, piirkond, emtak,
                                  rev_min, rev_max, emp_min, emp_max,
                                  profit_min, profit_max,
                                  limit, offset, sort_key, sort_dir)
    return jsonify({"total": total, "rows": rows})


@app.route("/api/companies")
def api_companies():
    return jsonify(load_companies())


@app.route("/api/companies/<int:row_num>", methods=["PATCH"])
def api_update(row_num: int):
    data = request.json
    for key, col_name in FIELD_MAP.items():
        if key in data:
            val = data[key]
            if key == 'score':
                try:
                    val = int(val) if val not in (None, '', 'Not set') else None
                except (ValueError, TypeError):
                    val = None
            save_field(row_num, col_name, val if val else None)
    return jsonify({"ok": True})


@app.route("/api/statuses")
def api_statuses():
    return jsonify(STATUSES)


# Build cache at import time (runs in both main process and Flask reloader child)
build_cache()

if __name__ == "__main__":
    app.run(debug=True, port=5000)
