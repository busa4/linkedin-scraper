"""
LinkedIn Company Search Scraper
Exports search results to Excel. Requires manual login on first run.
"""

import time
import random
from pathlib import Path
from datetime import datetime

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


SEARCH_URL = (
    "https://www.linkedin.com/search/results/companies/"
    "?origin=FACETED_SEARCH"
    "&companyHqGeo=%5B%22102974008%22%5D"
    "&companySize=%5B%22C%22%2C%22D%22%5D"
)

OUTPUT_FILE = (
    Path(__file__).parent
    / "good_result"
    / f"linkedin_companies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# Selectors based on actual page structure (role/aria attrs — stable across redesigns)
CARD_SELECTOR = "[role='listitem'][aria-labelledby]"
NEXT_BTN_SELECTOR = "[data-testid='pagination-controls-next-button-visible']"

BASE_DELAY = 4.0  # seconds between pages


def human_delay(base: float = BASE_DELAY) -> None:
    time.sleep(base * (0.7 + random.random() * 0.6))


def extract_companies(page) -> list[dict]:
    """Extract all company cards visible on the current page."""
    companies = []

    # Wait for at least one card to appear
    try:
        page.wait_for_selector(CARD_SELECTOR, timeout=15_000)
    except PlaywrightTimeout:
        print("  Warning: no result cards found on this page.")
        return companies

    cards = page.query_selector_all(CARD_SELECTOR)
    print(f"  Found {len(cards)} cards.")

    for card in cards:
        try:
            # Company name + URL: first anchor pointing to a /company/ path
            link_el = card.query_selector("a[href*='/company/']")
            if not link_el:
                continue

            name = link_el.inner_text().strip()
            url = link_el.get_attribute("href") or ""
            url = url.split("?")[0]  # drop tracking params

            # All <p> elements in the card — first one has the name link,
            # remaining ones carry industry, location, description
            all_p = card.query_selector_all("p")
            meta_texts = []
            for p in all_p:
                text = p.inner_text().strip()
                # Skip the paragraph that IS the company name
                if text and text != name:
                    meta_texts.append(text)

            industry = meta_texts[0] if len(meta_texts) > 0 else ""
            location = meta_texts[1] if len(meta_texts) > 1 else ""
            description = meta_texts[2] if len(meta_texts) > 2 else ""

            if name:
                companies.append(
                    {
                        "Name": name,
                        "LinkedIn URL": url,
                        "Industry": industry,
                        "Location": location,
                        "Description": description,
                    }
                )
        except Exception as exc:
            print(f"  Skipping card: {exc}")

    return companies


def scrape_all_pages(page) -> list[dict]:
    """Navigate through pages using the Next button and collect all companies."""
    all_companies: list[dict] = []
    page_num = 1

    while True:
        print(f"\nScraping page {page_num}…")
        human_delay(2)

        companies = extract_companies(page)
        all_companies.extend(companies)
        print(f"  Running total: {len(all_companies)} companies")

        # Check if Next button exists
        next_btn = page.query_selector(NEXT_BTN_SELECTOR)
        if not next_btn:
            print("  No Next button found — reached the last page.")
            break

        print("  Clicking Next…")
        next_btn.click()
        page_num += 1

        # Wait for the page to update (new cards load)
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
        except PlaywrightTimeout:
            pass
        human_delay()

    return all_companies


def save_to_excel(companies: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn Companies"

    headers = ["#", "Name", "LinkedIn URL", "Industry", "Location", "Description"]
    header_fill = PatternFill("solid", fgColor="0A66C2")  # LinkedIn blue
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, company in enumerate(companies, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=company["Name"])

        url = company["LinkedIn URL"]
        if url:
            cell = ws.cell(row=row_idx, column=3, value=url)
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
        else:
            ws.cell(row=row_idx, column=3, value="")

        ws.cell(row=row_idx, column=4, value=company["Industry"])
        ws.cell(row=row_idx, column=5, value=company["Location"])
        ws.cell(row=row_idx, column=6, value=company["Description"])

        if row_idx % 2 == 0:
            row_fill = PatternFill("solid", fgColor="EBF3FB")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = row_fill

    col_widths = [5, 35, 50, 30, 25, 60]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main():
    with sync_playwright() as p:
        print("Launching browser…")
        browser = p.chromium.launch(
            headless=False,
            args=["--start-maximized"],
        )
        context = browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()

        # Step 1: Manual login
        print("\nOpening LinkedIn login page.")
        print("Please log in in the browser window, then press Enter here.\n")
        page.goto("https://www.linkedin.com/login", wait_until="domcontentloaded")
        input(">>> Press Enter after you have logged in: ")

        # Step 2: Navigate to search
        print("\nNavigating to search results…")
        page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=30_000)
        human_delay(3)

        # Step 3: Scrape all pages
        all_companies = scrape_all_pages(page)

        # Step 4: Export
        print(f"\nSaving {len(all_companies)} companies to:\n  {OUTPUT_FILE}")
        save_to_excel(all_companies, OUTPUT_FILE)
        print("Done!")

        browser.close()


if __name__ == "__main__":
    main()
